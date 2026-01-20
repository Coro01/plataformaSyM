import csv
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum
from django.db import IntegrityError, transaction
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.views import LoginView
from django.http import JsonResponse, HttpResponse
from django.urls import reverse_lazy, reverse 
from .models import RegistroJornada, PerfilEmpleado, SolicitudLibre
from .forms import SolicitudLibreForm, EmpleadoLoginForm, UploadFileForm, RegistroJornadaForm 
from datetime import datetime, time, timedelta, date
from django.utils import timezone
from calendar import month_name
from django.contrib.auth.models import User
import re 
from functools import wraps
from django.shortcuts import redirect
from django.contrib import messages
from django.http import Http404
from django.shortcuts import get_object_or_404, redirect
from django.db import transaction
# Asegúrate de importar esto de tus utilidades
from .models import SolicitudLibre
from django.contrib.auth.decorators import login_required, user_passes_test
from datetime import timedelta # Asegúrate de tener esta importación
from openpyxl import load_workbook 
from decimal import Decimal
from django.urls import reverse # Asegúrate de importar reverse
from django.db.models import Q 
from openpyxl.utils.exceptions import InvalidFileException # Importación necesaria para manejar errores de archivo
# =================================================================
# 1. REGLAS DE NEGOCIO Y LÓGICA CENTRAL
# =================================================================

# --- REGLAS DE NEGOCIO GLOBALES ---
HORA_INICIO_OFICIAL = time(7, 0) # 07:00:00 (Hora antes de la cual no se suman horas)
DURACION_ALMUERZO = timedelta(minutes=30) # 30 minutos de descuento automático
JORNADA_ESTANDAR = timedelta(hours=9, minutes=30) # 9 horas y 30 minutos de jornada

# --- Funciones de Comprobación de Permisos ---
def is_staff_check(user):
    """Verifica si el usuario es un administrador o staff."""
    return user.is_active and user.is_staff

def requiere_nivel(nivel_minimo):
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                messages.error(request, '❌ Debes iniciar sesión.')
                return redirect('control_horas:login')
            
            try:
                # Usar el nombre correcto del atributo OneToOne
                perfil = request.user.perfilempleado 
            except AttributeError: # O PerfilEmpleado.DoesNotExist si importas el modelo
                messages.error(request, '❌ No tienes un perfil configurado.')
                # Redirigir a una ruta que SI existe en urls.py
                return redirect('control_horas:main_dashboard') 
            
            if not perfil.tiene_acceso(nivel_minimo):
                # Mensaje amigable en lugar de error 404
                messages.error(request, f'🚫 Acceso denegado: Tu nivel actual ({perfil.get_nivel_display()}) no permite ingresar a esta sección.')
                
                # Redirigir a la página de donde venía (o al inicio si no hay anterior)
                return redirect(request.META.get('HTTP_REFERER', 'control_horas:main_dashboard'))
            
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator

# --- Funciones de Cálculo Central ---
# --- Función Auxiliar (Colocar al inicio de views.py con otras funciones) ---
@login_required
def main_dashboard_view(request):
    """Vista para el Dashboard Principal (Selector de Módulos)."""
    return render(request, 'control_horas/main_dashboard.html')

def save_incomplete_jornada(empleado_user_obj, fecha, entrada):
    """
    Guarda un registro de jornada incompleto (solo entrada).
    Lo marca como 'salida_forzada=True' y establece la salida por defecto
    para permitir que el registro exista y pueda ser corregido.
    """
    # Usamos una hora de salida por defecto (p. ej., 17:00) 
    # Las horas netas y extras se dejan en None/cero.
    SALIDA_POR_DEFECTO = time(hour=17, minute=0) 
    
    with transaction.atomic():
        RegistroJornada.objects.update_or_create(
            empleado=empleado_user_obj, 
            fecha=fecha,
            defaults={
                'entrada': entrada, 
                'salida': SALIDA_POR_DEFECTO, 
                'horas_netas': None, 
                'horas_extras': timedelta(hours=0),
                'salida_forzada': True, # Marcador CRÍTICO para indicar que necesita revisión
            }
        )

def calcular_horas_jornada(fecha: date, entrada: time, salida: time):
    """
    Calcula la duración neta de la jornada y las horas extras basándose en
    las reglas de negocio (HORA_INICIO_OFICIAL, DURACION_ALMUERZO, JORNADA_ESTANDAR).

    Args:
        fecha (date): Fecha del registro.
        entrada (time): Hora de entrada registrada.
        salida (time): Hora de salida registrada.

    Returns:
        tuple: (duracion_neta, horas_extras), ambos como timedelta.
    """
    # 1. Ajustar la hora de entrada si es anterior a la HORA_INICIO_OFICIAL
    # Esto asume que el campo 'entrada' del formulario ya está siendo validado
    # y es un objeto time válido.
    hora_inicio_efectiva = max(entrada, HORA_INICIO_OFICIAL)

    # 2. Calcular la duración bruta 
    dt_entrada = datetime.combine(fecha, hora_inicio_efectiva)
    dt_salida = datetime.combine(fecha, salida)
    
    # Manejar el caso de jornada nocturna (salida al día siguiente)
    if dt_salida < dt_entrada:
        dt_salida += timedelta(days=1)

    duracion_bruta = dt_salida - dt_entrada

    # 3. Calcular la duración neta descontando el almuerzo
    duracion_neta = duracion_bruta - DURACION_ALMUERZO
    
    if duracion_neta < timedelta(0):
        duracion_neta = timedelta(0)

    # 4. Calcular horas extras (saldo diario)
    horas_extras = duracion_neta - JORNADA_ESTANDAR
    
    return duracion_neta, horas_extras

def format_timedelta_to_hhmmss(td: timedelta) -> str:
    """Formatea un timedelta a una cadena HH:MM:SS, manejando valores negativos."""
    if not td:
        return "0:00:00"

    total_seconds = int(td.total_seconds())
    sign = '-' if total_seconds < 0 else ''
    abs_seconds = abs(total_seconds)

    hours = abs_seconds // 3600
    minutes = (abs_seconds % 3600) // 60
    seconds = abs_seconds % 60
    
    return f"{sign}{hours:01d}:{minutes:02d}:{seconds:02d}"

def obtener_saldo_actual(empleado):
    """
    Calcula el saldo total de horas compensatorias acumuladas (Ganadas - Usadas).
    
    El parámetro 'empleado' es el objeto User (user_instance) que pasas desde la vista.
    """
    
    # 1. Horas Extras Ganadas (acumuladas)
    total_extras_ganadas = RegistroJornada.objects.filter(
        empleado=empleado
    ).aggregate(Sum('horas_extras'))['horas_extras__sum'] or timedelta(hours=0)
    
    # 2. Horas Libres APROBADAS (utilizadas)
    total_horas_usadas = SolicitudLibre.objects.filter(
        empleado=empleado,
        estado='APROBADO' # Solo contamos las que ya fueron aprobadas
    ).aggregate(Sum('horas_solicitadas'))['horas_solicitadas__sum'] or timedelta(hours=0)
    
    # 3. Retorna el saldo total
    return total_extras_ganadas - total_horas_usadas

# --- FUNCIÓN AUXILIAR PARA CÁLCULO DE SALDO MENSUAL ---
def obtener_saldo_mensual(empleado, mes_num, anio):
    """Calcula el saldo total de horas extras acumuladas del empleado para un mes específico."""
    
    saldo_mensual = RegistroJornada.objects.filter(
        empleado=empleado,
        fecha__year=anio,
        fecha__month=mes_num
    ).aggregate(
        saldo_mensual=Sum('horas_extras')
    )['saldo_mensual'] or timedelta(0)
    
    return saldo_mensual


# --- FUNCIÓN AUXILIAR DE FORMATO DE TIEMPO (CORRECCIÓN CLAVE) ---
def format_timedelta_display(td: timedelta):
    """Convierte un timedelta a una cadena H:MM con signo para una mejor visualización."""
    if not td:
        return "0:00"
    
    total_seconds = int(td.total_seconds())
    sign = '-' if total_seconds < 0 else ''
    abs_seconds = abs(total_seconds)
    hours = abs_seconds // 3600
    minutes = (abs_seconds % 3600) // 60
    # Retorna un string simple: -H:MM o H:MM
    return f"{sign}{hours}:{minutes:02d}"

# --- FUNCIÓN AUXILIAR DE FORMATO DE TIEMPO (HH:MM:SS) ---
def format_timedelta_to_hhmmss(td):
    """Convierte un objeto timedelta a un string 'HH:MM:SS'. Útil para exportaciones."""
    if td is None:
        return '00:00:00'
    total_seconds = int(td.total_seconds())
    sign = '-' if total_seconds < 0 else ''
    total_seconds = abs(total_seconds)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"

# =================================================================
# 2. VISTAS DE AUTENTICACIÓN
# =================================================================

class EmpleadoLoginView(LoginView):
    """Vista de Login personalizada usando el formulario EmpleadoLoginForm."""
    template_name = 'control_horas/login.html'
    authentication_form = EmpleadoLoginForm

    def get_success_url(self):
        """Redirige al saldo de horas tras el login exitoso."""
        return reverse_lazy('control_horas:main_dashboard')

@login_required
def logout_confirm_view(request):
    """Muestra una pantalla de confirmación antes de hacer logout."""
    if request.method == 'POST':
        logout(request)
  
        return redirect('control_horas:login')
    return render(request, 'control_horas/logout_confirm.html')


# =================================================================
# 3. VISTAS PRINCIPALES Y DATOS (CRUD y API)
# =================================================================

@requiere_nivel(1)
def saldo_horas_view(request):
    """Muestra el saldo de horas acumuladas y mensuales de todos los empleados (si es staff) o solo el propio."""
    
    hoy = date.today()
    mes_actual_num = hoy.month
    año_actual = hoy.year
    nombre_mes = month_name[mes_actual_num].capitalize()
    
    contexto_empleados = []
    
    # Determinar qué empleados mostrar
    if request.user.is_staff:
        empleados_a_mostrar = PerfilEmpleado.objects.all().select_related('user')
    else:
        try:
            empleados_a_mostrar = [request.user.perfilempleado]
        except PerfilEmpleado.DoesNotExist:
            messages.error(request, "Error: Su cuenta de usuario no tiene un perfil de empleado asociado.")
            return redirect('control_horas:logout_confirm')

    for perfil in empleados_a_mostrar:
        user_instance = perfil.user
        
        # 1. CALCULAR SALDO TOTAL ACUMULADO
        total_acumulado_td = obtener_saldo_actual(user_instance)
        
        # 2. CALCULAR SALDO MENSUAL
        saldo_mensual_td = obtener_saldo_mensual(user_instance, mes_actual_num, año_actual)
        
        # 3. OBTENER Y FORMATEAR ÚLTIMOS REGISTROS
        registros_formateados = []
        # CORRECCIÓN: Filtramos por el objeto User que está en la FK
        registros_qs = RegistroJornada.objects.filter(empleado=user_instance).order_by('-fecha')[:5]
        
        for reg in registros_qs:
            registros_formateados.append({
                'fecha': reg.fecha,
                'entrada': reg.entrada,
                'salida': reg.salida,
                'id': reg.id,
                'horas_extras_display': format_timedelta_display(reg.horas_extras),
            })
            
        # 4. Construir el objeto de contexto
        contexto_empleados.append({
            'nombre': user_instance.username,
            'saldo_total': format_timedelta_display(total_acumulado_td),
            'saldo_mensual': format_timedelta_display(saldo_mensual_td),
            'registros': registros_formateados,
        })
        
    context = {
        'contexto_empleados': contexto_empleados,
        'mes_actual': nombre_mes,
        'messages': messages.get_messages(request),
    }
    
    return render(request, 'control_horas/saldo_horas.html', context)

@requiere_nivel(1)
def calendario_horas_view(request):
    """
    Muestra el calendario enfocado en registros de jornada (entradas/salidas/saldo).
    """
    user = request.user
    if user.is_staff:
        empleados_a_filtrar = PerfilEmpleado.objects.all().select_related('user').order_by('user__username')
    else:
        try:
            empleados_a_filtrar = [user.perfilempleado]
        except PerfilEmpleado.DoesNotExist:
            empleados_a_filtrar = []

    context = {
        'empleados': empleados_a_filtrar,
        'is_staff': user.is_staff,
        'titulo_calendario': 'Calendario de Jornada y Horas Extra',
        'api_endpoint_name': 'control_horas:journal_data_api',
        'calendar_type': 'horas'
    }
    return render(request, 'control_horas/calendario_horas.html', context)


@requiere_nivel(1)
def calendario_solicitudes_view(request):
    """
    Muestra el calendario enfocado en solicitudes de días libres.
    """
    user = request.user
    if user.is_staff:
        empleados_a_filtrar = PerfilEmpleado.objects.all().select_related('user').order_by('user__username')
    else:
        try:
            empleados_a_filtrar = [user.perfilempleado]
        except PerfilEmpleado.DoesNotExist:
            empleados_a_filtrar = []

    context = {
        'empleados': empleados_a_filtrar,
        'is_staff': user.is_staff,
        'titulo_calendario': 'Solicitudes de Días Libres',
        'api_endpoint_name': 'control_horas:journal_data_api',
        'calendar_type': 'solicitudes'
    }
    return render(request, 'control_horas/calendario_solicitudes.html', context)


# =================================================================
# 3. API ENDPOINT UNIFICADO (CORREGIDO)
# =================================================================
@requiere_nivel(1)
def journal_data_api(request):
    user = request.user
    # Usamos '' como valor por defecto para manejar correctamente 'Todos los funcionarios'
    empleado_username = request.GET.get('empleado', '') 
    tipo_dato = request.GET.get('tipo') 
    
    if not tipo_dato or tipo_dato not in ['horas', 'solicitudes']:
        return JsonResponse({'error': 'Tipo de dato (tipo) es requerido o inválido.'}, status=400)

    # 1. FILTRO BASE: Rango de Fechas
    q_filter = Q()
    
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    
    try:
        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            # Filtra por el campo de fecha correcto según el tipo de dato
            date_field = 'fecha' if tipo_dato == 'horas' else 'fecha_libre'
            q_filter &= Q(**{f'{date_field}__gte': fecha_inicio})
        
        if fecha_fin_str:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            date_field = 'fecha' if tipo_dato == 'horas' else 'fecha_libre'
            q_filter &= Q(**{f'{date_field}__lte': fecha_fin})
            
    except ValueError:
        # Si el formato de fecha es incorrecto, simplemente no aplica el filtro de fecha
        pass 
        
    data = []
    
    # ----------------------------------------------------------------------
    # LÓGICA PARA JORNADAS DE TRABAJO (TIPO: HORAS)
    # ----------------------------------------------------------------------
    if tipo_dato == 'horas':
        from .models import RegistroJornada
        
        # 2. FILTRO DE EMPLEADO para HORAS
        horas_empleado_filter = Q()
        if user.is_staff:
            # Si es staff y se ha filtrado a un empleado específico:
            if empleado_username: 
                horas_empleado_filter &= Q(empleado__username__iexact=empleado_username)
            # Si no hay empleado_username (''), el filtro queda Q(), mostrando todos.
        else:
            # Si NO es staff, solo ve sus propias jornadas
            horas_empleado_filter &= Q(empleado=user) 

        # Se aplica el filtro de fechas (q_filter) Y el de empleado
        jornadas_qs = RegistroJornada.objects.filter(q_filter & horas_empleado_filter).select_related('empleado').order_by('fecha')
        
        # ... (rest of the processing logic for 'horas' is correct and maintained)
        for jornada in jornadas_qs:
            empleado_nombre = jornada.empleado.username.capitalize() 
            
            url_edicion = None
            # Permitir edición solo a Staff o al propio empleado
            if user.is_staff or user.username == jornada.empleado.username: 
                url_edicion = reverse('control_horas:editar_registro', args=[jornada.pk])
            
            # --- CORRECCIÓN DE DATOS PARA INCOMPLETO/COMPLETO ---
            
            if jornada.salida_forzada or jornada.horas_netas is None:
                salida_fmt = jornada.salida.strftime('%H:%M') if jornada.salida else 'N/A'
                
                data.append({
                    'id': jornada.pk,
                    'title': f'⚠️ Error: Entrada {jornada.entrada.strftime("%H:%M")}',
                    'start': jornada.fecha.isoformat(),
                    'allDay': True,
                    'backgroundColor': '#f97316', 
                    'borderColor': '#f97316',
                    'url': url_edicion,
                    'extendedProps': {
                        'tipo': 'incompleto', 
                        'empleado': empleado_nombre,
                        'entrada': jornada.entrada.strftime('%H:%M'),
                        'salida': salida_fmt, 
                        'horas_netas': 'N/A',
                        'horas_extras': 'N/A', 
                        'mensaje': 'Registro Incompleto. Requiere corrección.',
                    },
                    'className': 'fc-event-jornada-base'
                })
            
            else: 
                try:
                    # Se asume que format_timedelta_display está definido. Si no, se usa str().
                    saldo_display = format_timedelta_display(jornada.horas_extras)
                    horas_netas_display = format_timedelta_display(jornada.horas_netas)
                except NameError:
                    saldo_display = str(jornada.horas_extras)
                    horas_netas_display = str(jornada.horas_netas)
                
                if jornada.horas_extras and jornada.horas_extras > timedelta(minutes=0):
                    bg_color = '#10b981'
                    title = f'✅ {horas_netas_display} (+{saldo_display} Extra)'
                else:
                    bg_color = '#60a5fa'
                    title = f'✔️ {horas_netas_display}'

                data.append({
                    'id': jornada.pk,
                    'title': title, 
                    'start': jornada.fecha.isoformat(),
                    'allDay': True, 
                    'backgroundColor': bg_color,
                    'borderColor': bg_color,
                    'url': url_edicion,
                    'extendedProps': {
                        'tipo': 'completo', 
                        'empleado': empleado_nombre,
                        'entrada': jornada.entrada.strftime('%H:%M'),
                        'salida': jornada.salida.strftime('%H:%M'),
                        'horas_netas': horas_netas_display,
                        'horas_extras': saldo_display, 
                    },
                    'className': 'fc-event-jornada-base' 
                })


    # ----------------------------------------------------------------------
    # LÓGICA PARA SOLICITUDES (TIPO: SOLICITUDES) - REFORZADO
    # ----------------------------------------------------------------------
    elif tipo_dato == 'solicitudes':
        from .models import SolicitudLibre 
        
        # 2. FILTRO DE EMPLEADO para SOLICITUDES (Aplicando restricción para no-staff)
        solicitudes_empleado_filter = Q()
        
        if user.is_staff:
            # Staff: ve todas o las del empleado filtrado.
            if empleado_username:
                solicitudes_empleado_filter &= Q(empleado__username=empleado_username) 
        else:
            # No-Staff: solo ve sus propias solicitudes.
            solicitudes_empleado_filter &= Q(empleado=user)


        # 3. EJECUTAR QUERYSET CON AMBOS FILTROS
        solicitudes = SolicitudLibre.objects.filter(q_filter & solicitudes_empleado_filter).select_related('empleado').order_by('fecha_libre')
        
        # 4. PROCESAR DATOS
        for sol in solicitudes:
            empleado_nombre = sol.empleado.username.capitalize()

            try:
                horas_solicitadas_display = format_timedelta_display(sol.horas_solicitadas)
            except NameError:
                horas_solicitadas_display = str(sol.horas_solicitadas)

            color_map = {'APROBADO': '#10b981', 'PENDIENTE': '#f59e0b', 'RECHAZADO': '#dc3545'}
            class_map = {'APROBADO': 'fc-event-day_libre-aprobado', 'PENDIENTE': 'fc-event-day_libre-pendiente', 'RECHAZADO': 'fc-event-day_libre-rechazado'}
            
            data.append({
                'id': f'sol-{sol.pk}',
                'title': f'{empleado_nombre} ({sol.estado})',
                'start': sol.fecha_libre.isoformat(),
                'allDay': True,
                'backgroundColor': color_map.get(sol.estado, '#6c757d'),
                'borderColor': color_map.get(sol.estado, '#6c757d'),
                'extendedProps': {
                    'empleado': empleado_nombre,
                    'estado': sol.estado,
                    'horas_solicitadas': horas_solicitadas_display,
                    'motivo': sol.motivo or 'No especificado', 
                    'tipo': 'dia_libre',
                },
                'className': class_map.get(sol.estado, 'fc-event-other')
            })

    return JsonResponse(data, safe=False)
@requiere_nivel(5)
def editar_registro_jornada_view(request, registro_id):
    
    # 1. Obtener la instancia del registro
    registro = get_object_or_404(RegistroJornada, pk=registro_id) 
    
    empleado_actual = request.user
    
    # 2. Restricción de seguridad
    if registro.empleado != empleado_actual and not empleado_actual.is_staff:
        messages.error(request, "❌ No tiene permiso para editar este registro.")
        return redirect('control_horas:calendario_horas') 
    
    # 3. Lógica del formulario
    if request.method == 'POST':
        form = RegistroJornadaForm(request.POST, instance=registro, user=request.user)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    registro_actualizado = form.save(commit=False)
                    
                    entrada = form.cleaned_data.get('entrada')
                    salida = form.cleaned_data.get('salida')
                    
                    # --- CÁLCULO DE HORAS ---
                    if entrada and salida:
                        # Se usa la fecha del registro, y las horas limpias del formulario
                        horas_netas, horas_extras = calcular_horas_jornada(registro_actualizado.fecha, entrada, salida) 
                        
                        registro_actualizado.horas_netas = horas_netas
                        registro_actualizado.horas_extras = horas_extras
                        registro_actualizado.salida_forzada = False 
                    else:
                        # Si faltan datos de tiempo
                        registro_actualizado.horas_netas = None
                        registro_actualizado.horas_extras = timedelta(0)
                        registro_actualizado.salida_forzada = True
                    
                    # 4. Guardar el registro
                    registro_actualizado.save()
                    
                    # 5. Recalcular y actualizar el saldo global del PerfilEmpleado
                    empleado_user_obj = registro_actualizado.empleado
                    nuevo_saldo = obtener_saldo_actual(empleado_user_obj)
                    
                    # Se asume que PerfilEmpleado siempre existe
                    empleado_perfil_obj = get_object_or_404(PerfilEmpleado, user=empleado_user_obj)
                    empleado_perfil_obj.saldo_horas_comp = nuevo_saldo
                    empleado_perfil_obj.save()
                    
                    messages.success(request, f"✅ Registro #{registro.pk} editado correctamente. Saldo actualizado: {format_timedelta_display(nuevo_saldo)}.")
                    return redirect('control_horas:calendario_horas') 
                
            except Exception as e:
                messages.error(request, f"Ocurrió un error al guardar el registro: {e}")
        else:
            messages.error(request, "Error de validación. Revise los campos marcados.")
    else:
        # Petición GET: Inicializa el formulario con los datos existentes
        form = RegistroJornadaForm(instance=registro, user=request.user)

    context = {
        'registro': registro,
        'form': form,
    }
    return render(request, 'control_horas/editar_registro.html', context)
# =================================================================
# 4. VISTAS DE SOLICITUDES DE DÍA LIBRE
# =================================================================


from datetime import timedelta # Asegúrate de importar esto

# control_horas/views.py (función solicitar_dia_libre_view)

@requiere_nivel(1)
def solicitar_dia_libre_view(request):
    empleado = request.user
    
    # Pre-cálculos necesarios para el contexto (GET y POST)
    saldo_total = obtener_saldo_actual(empleado)
    saldo_display = format_timedelta_display(saldo_total)
    solicitudes_recientes = SolicitudLibre.objects.filter(empleado=empleado).order_by('-fecha_solicitud')[:5]
    
    form = SolicitudLibreForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        
        solicitud = form.save(commit=False)
        solicitud.empleado = empleado
        
        # 1. ADVERTENCIA: Verifica si el saldo es negativo.
        if saldo_total < timedelta(0):
            messages.warning(
                request, 
                f"ADVERTENCIA: Tienes un saldo NEGATIVO ({saldo_display}). Tu solicitud se enviará para aprobación especial."
            )
        
        # 2. PROCESAMIENTO: La solicitud siempre se guarda y redirige, independientemente del saldo.
        try:
            solicitud.save()
            messages.success(
                request, 
                f"Solicitud de {format_timedelta_display(solicitud.horas_solicitadas)} para el {solicitud.fecha_libre} enviada. Esperando aprobación."
            )
            return redirect('control_horas:saldo_horas') 
            
        except Exception as e:
            messages.error(request, f"Error interno al guardar la solicitud: {e}")
            
    # ✅ LÓGICA DE CONTADOR CORREGIDA: Filtra por usuario a menos que sea administrador
    pendientes_qs = SolicitudLibre.objects.filter(estado='PENDIENTE')
    
    if not request.user.is_staff:
        # Si NO es administrador, solo se muestran sus propias solicitudes pendientes
        pendientes_qs = pendientes_qs.filter(empleado=empleado) 
        
    pendientes_count = pendientes_qs.count()
            
    # Contexto final para GET o POST fallido
    context = {
        'form': form,
        'saldo_total': saldo_total, 
        'saldo_display': saldo_display, 
        'solicitudes_recientes': solicitudes_recientes,
        'pendientes_count': pendientes_count, # <-- Esta variable ahora contiene el valor correcto
    }
    
    return render(request, 'control_horas/solicitar_dia_libre.html', context)
@requiere_nivel(5)
def gestion_solicitudes_view(request):
    """
    Vista de gestión para administradores: ver solicitudes pendientes y pasadas.
    """
    # Solicitudes pendientes (más importantes)
    pendientes = SolicitudLibre.objects.filter(estado='PENDIENTE').order_by('fecha_solicitud')
    
    # Solicitudes recientes (últimas 20 aprobadas/rechazadas)
    recientes = SolicitudLibre.objects.exclude(estado='PENDIENTE').order_by('-fecha_solicitud')[:20]
    
    # Obtener el saldo de cada empleado con solicitudes pendientes
    empleados_con_solicitudes = pendientes.values_list('empleado', flat=True).distinct()
    saldos_empleados = {
        User.objects.get(pk=user_id).username: format_timedelta_display(obtener_saldo_actual(User.objects.get(pk=user_id)))
        for user_id in empleados_con_solicitudes
    }
    
    context = {
        'pendientes': pendientes,
        'recientes': recientes,
        'saldos_empleados': saldos_empleados,
    }
    return render(request, 'control_horas/gestion_solicitudes.html', context)

@requiere_nivel(5)
@transaction.atomic
def aprobar_rechazar_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudLibre, pk=solicitud_id)
    
    if solicitud.estado != 'PENDIENTE':
        messages.error(request, f"❌ La solicitud ID {solicitud_id} ya ha sido '{solicitud.estado}'. No se puede modificar.")
        return redirect('control_horas:gestion_solicitudes')

    # 'empleado' es el objeto User (<User: Marcos>)
    empleado = solicitud.empleado 

    try:
        # Obtenemos el PerfilEmpleado
        perfil_empleado = empleado.perfilempleado 
    except PerfilEmpleado.DoesNotExist:
        messages.error(request, f"❌ ERROR: El usuario {empleado.username} no tiene PerfilEmpleado asociado.")
        return redirect('control_horas:gestion_solicitudes')
    except AttributeError:
        messages.error(request, f"❌ ERROR: La solicitud no tiene empleado válido.")
        return redirect('control_horas:gestion_solicitudes')


    accion = request.POST.get('accion') 
    accion_limpia = accion.lower() if accion else None

    if accion_limpia == 'aprobado':
        
        # Se calcula el saldo, pero SOLO para el mensaje final
        saldo_actual = obtener_saldo_actual(empleado) 
        
        # ❌ CONDICIÓN ELIMINADA: NO hay verificación de saldo insuficiente ❌
        
        # Descuento de Horas (Proceder siempre con el descuento)
        perfil_empleado.saldo_horas_comp -= solicitud.horas_solicitadas
        perfil_empleado.save() 
        
        # Finaliza la solicitud
        solicitud.estado = 'APROBADO'
        solicitud.save()
        
        # El saldo mostrado ahora puede ser negativo
        messages.success(request, f"✅ Solicitud de {empleado.username} para el {solicitud.fecha_libre} ha sido APROBADA. Nuevo saldo: {format_timedelta_display(perfil_empleado.saldo_horas_comp)}.")
        
    elif accion_limpia == 'rechazado':
        solicitud.estado = 'RECHAZADO'
        solicitud.save()
        messages.info(request, f"ℹ️ Solicitud de {empleado.username} para el {solicitud.fecha_libre} ha sido RECHAZADA.")
    else:
        messages.error(request, f"❌ Acción no válida recibida.")
        
    return redirect('control_horas:gestion_solicitudes')
@requiere_nivel(5)
def upload_excel_view(request):
    """
    Maneja la carga de archivos de jornada. Usa openpyxl para XLSX/XLSM
    y la lógica de regex/CSV como fallback para otros formatos.
    """
    
    # Declaración de variables globales requeridas por utilidades externas
    global HORA_INICIO_OFICIAL, DURACION_ALMUERZO, JORNADA_ESTANDAR 
    
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        registros_actualizados = 0
        errores = []

        # Determinar si es XLSX (opción B) o CSV (fallback)
        file_extension = archivo.name.split('.')[-1].lower()
        is_excel = file_extension in ['xlsx', 'xlsm']

        try:
            file_data_str = ""
            nro_afiliacion = None

            if is_excel:
                # --- 1. PROCESAMIENTO DE ARCHIVOS XLSX (openpyxl) ---
                workbook = load_workbook(archivo)
                sheet = workbook.active
                
                # Extracción directa de la celda E3
                nro_afiliacion_raw = sheet['E3'].value
                
                if nro_afiliacion_raw:
                    nro_afiliacion = str(nro_afiliacion_raw).strip() 
                
                # Convertir la hoja activa en formato CSV en memoria
                buffer = io.StringIO()
                writer = csv.writer(buffer)

                for row in sheet.iter_rows():
                    writer.writerow([cell.value if cell.value is not None else '' for cell in row])
                
                file_data_str = buffer.getvalue()
                
            else:
                # --- 1. PROCESAMIENTO DE ARCHIVOS CSV (Fallback/Original) ---
                file_data_str = archivo.read().decode('latin-1')
                file_data_str = file_data_str.replace('\r\n', '\n').replace('\r', '\n')
                
                # Extracción con REGEX para CSV
                pattern = r'N.*?AFILIAC.*?(\d+)'
                nro_afiliacion = None
                match = re.search(pattern, file_data_str, re.IGNORECASE | re.DOTALL) 
                if match:
                    nro_afiliacion = match.group(1).strip()
            
            # --- 2. VALIDACIÓN Y BÚSQUEDA POR NÚMERO ---
            if not nro_afiliacion:
                # Mensaje de error si no se encuentra el Nº de Afiliación
                primeras_lineas = '\n'.join(file_data_str.split('\n')[:5]).replace('"', '')
                messages.error(request, f"Error: ¡Fallo crítico! No se pudo extraer el **Número de Afiliación (Cédula)**. Primeras líneas para diagnóstico:\n---\n{primeras_lineas}\n---")
                return redirect('control_horas:saldo_horas')

            # Buscar por el Número de Afiliación en la base de datos
            # Nota: Asume que PerfilEmpleado está correctamente importado
            empleado_perfil_obj = PerfilEmpleado.objects.filter(nro_afiliacion=nro_afiliacion).first()
            if not empleado_perfil_obj:
                messages.error(request, f"Error: No se encontró un perfil con el Número de Afiliación: **{nro_afiliacion}**.")
                return redirect('control_horas:saldo_horas')

            empleado_user_obj = empleado_perfil_obj.user
            
            # --- 3. PROCESAMIENTO DE JORNADAS ---
            csv_reader = csv.reader(io.StringIO(file_data_str), delimiter=',', quotechar='"')
            
            fecha_actual = None
            registro_jornada = {}
            
            # Mapeo de meses en español a inglés para el parseo
            month_map = {
                'octubre': 'October', 'noviembre': 'November', 'diciembre': 'December',
                'enero': 'January', 'febrero': 'February', 'marzo': 'March',
                'abril': 'April', 'mayo': 'May', 'junio': 'June',
                'julio': 'July', 'agosto': 'August', 'septiembre': 'September',
            }
            
            event_map = {
                'inicio de jornada': 'entrada',
                'finaliza la jornada': 'salida',
            }

            with transaction.atomic():
                
                for row_num, row in enumerate(csv_reader):
                    # Saltar encabezados, resúmenes y pausas de comida
                    if len(row) > 0 and 'RESUMEN' in row[0].upper(): break 
                    if not row or not row[0] or 'total tiempo' in row[0].lower() or 'tiempo total' in row[0].lower(): continue
                    if len(row) > 1 and row[1].strip().lower() in ['pausa de jornada (comida)', 'reanuda la jornada']: continue 

                    # A. DETECTAR Y PARSEAR LA FECHA (Inicio de un nuevo día)
                    if ',' in row[0] and any(m in row[0].lower() for m in month_map.keys()): 
                        try:
                            # Parsear la fecha
                            raw_date_part = re.sub(r'^\w+,\s*|\s*\d{2}:\d{2}$', '', row[0].strip().strip('"')).strip()
                            partes_fecha_en = raw_date_part
                            for es, en in month_map.items():
                                partes_fecha_en = partes_fecha_en.replace(es, en)
                                
                            nueva_fecha = datetime.strptime(partes_fecha_en, '%d %B %Y').date()
                            
                            # --- LÓGICA DE VALIDACIÓN DE JORNADA ABIERTA ---
                            if fecha_actual and registro_jornada.get('entrada') and not registro_jornada.get('salida'):
                                errores.append(f"⚠️ REGISTRO INCOMPLETO DETECTADO: {fecha_actual.isoformat()}")
                                save_incomplete_jornada(empleado_user_obj, fecha_actual, registro_jornada['entrada'])
                                
                            # Asignar la nueva fecha y resetear el seguimiento de la jornada
                            fecha_actual = nueva_fecha
                            registro_jornada = {'entrada': None, 'salida': None}
                            continue
                        except Exception:
                            # Ignorar líneas que parecen fechas pero no se pueden parsear
                            pass 

                    # B. DETECTAR Y ASIGNAR HORA (Entrada/Salida)
                    if fecha_actual and len(row) > 1 and ':' in row[0] and row[1].strip().lower() in event_map:
                        try:
                            hora_str_clean = row[0].split(' ')[0].strip()
                            tipo_evento = row[1].strip().lower()
                            
                            hora_evento = datetime.strptime(hora_str_clean, '%H:%M').time()
                            campo = event_map[tipo_evento]
                            registro_jornada[campo] = hora_evento
                            
                            # Si se detecta la SALIDA y tenemos ENTRADA, calcular y guardar
                            if campo == 'salida' and registro_jornada.get('entrada'):
                                
                                # --- REGLAS DE CÁLCULO ---
                                # Nota: Asume que calcular_horas_jornada está disponible
                                duracion_neta, horas_extras = calcular_horas_jornada(
                                    fecha=fecha_actual,
                                    entrada=registro_jornada['entrada'],
                                    salida=registro_jornada['salida']
                                )
                                
                                # Crear o Actualizar el registro
                                RegistroJornada.objects.update_or_create(
                                    empleado=empleado_user_obj, 
                                    fecha=fecha_actual,
                                    defaults={
                                        'entrada': registro_jornada['entrada'], 
                                        'salida': registro_jornada['salida'],
                                        'horas_netas': duracion_neta,
                                        'horas_extras': horas_extras,
                                        'salida_forzada': False, # Registro completo
                                    }
                                )
                                registros_actualizados += 1
                                # Reiniciar el seguimiento
                                fecha_actual = None 
                                registro_jornada = {}
                                
                        except Exception as e:
                            # Captura errores en el parseo o cálculo
                            errores.append(f"Error procesando la jornada de **{fecha_actual}** (Fila {row_num + 1}). Detalles: {e}")
                            
                            # Si hubo un error en el cálculo, intentamos guardar al menos la entrada
                            if fecha_actual and registro_jornada.get('entrada'):
                                save_incomplete_jornada(empleado_user_obj, fecha_actual, registro_jornada['entrada'])
                                
                            fecha_actual = None 
                            registro_jornada = {}

                if errores:
                    messages.warning(request, f"¡Atención! Se cargaron {registros_actualizados} registros, pero hubo registros incompletos en {len(errores)} días. Revise los mensajes de error individuales.")
                    # Los errores se muestran uno por uno
                    for err in errores:
                        messages.error(request, err)
                else:
                    messages.success(request, f"¡Procesamiento finalizado! Se cargaron/actualizaron {registros_actualizados} registros para {empleado_perfil_obj.user.username.capitalize()}.")
            
        except InvalidFileException:
            messages.error(request, "Error de procesamiento: El archivo subido no es un archivo **Excel (.xlsx)** válido o está corrupto. Intente guardarlo de nuevo.")
        except Exception as e:
            # Captura errores generales
            messages.error(request, f"Error al procesar el archivo. Asegúrese de que no esté protegido por contraseña o corrupto. Detalles: {e}")

        return redirect('control_horas:saldo_horas')
    
    # --- LÓGICA DE GET: Muestra el formulario ---
    form = UploadFileForm()
    
    # IMPORTANTE: Eliminamos la lógica de recolectar mensajes manualmente. 
    # El template base (o upload_excel.html) ya itera sobre 'messages'.
    context = {
        'form': form,
        # Si el template upload_excel.html tiene un bucle 'if error',
        # puedes eliminarlo y usar solo el bucle 'if messages'.
    }
    return render(request, 'control_horas/upload_excel.html', context)

@requiere_nivel(1)
def calculadora_view(request):
    """
    Muestra una calculadora de jornada simple que replica la lógica de negocio.
    """
    resultado = None
    
    if request.method == 'POST':
        try:
            fecha_str = request.POST.get('fecha', timezone.now().date().isoformat())
            entrada_str = request.POST.get('entrada')
            salida_str = request.POST.get('salida')
            
            # Conversión de strings a objetos date/time
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            entrada = datetime.strptime(entrada_str, '%H:%M').time()
            salida = datetime.strptime(salida_str, '%H:%M').time()
            
            # Calcular
            duracion_neta, horas_extras = calcular_horas_jornada(fecha, entrada, salida)
            
            # Usar la función auxiliar global para formatear timedelta
            duracion_bruta = datetime.combine(fecha, salida) - datetime.combine(fecha, entrada)

            resultado = {
                'entrada': entrada.strftime("%H:%M"),
                'salida': salida.strftime("%H:%M"),
                'fecha': fecha.strftime("%d/%m/%Y"),
                'duracion_neta': format_timedelta_to_hhmmss(duracion_neta),
                'horas_extras': format_timedelta_to_hhmmss(horas_extras),
                'jornada_estandar': format_timedelta_to_hhmmss(JORNADA_ESTANDAR),
                'descuento_almuerzo': format_timedelta_to_hhmmss(DURACION_ALMUERZO) if duracion_bruta >= DURACION_ALMUERZO else '00:00:00',
            }
            
        except ValueError as e:
            messages.error(request, f"Error en el formato de datos. Asegúrese de que las horas estén en HH:MM y la fecha en AAAA-MM-DD. Detalle: {e}")
        except Exception as e:
            messages.error(request, f"Ocurrió un error inesperado al calcular. Detalle: {e}")

    # Valores por defecto para el formulario GET o en caso de error
    context = {
        'resultado': resultado,
        'fecha_default': timezone.now().date().isoformat(),
        'entrada_default': '08:00',
        'salida_default': '18:00',
        'hora_inicio_oficial': HORA_INICIO_OFICIAL.strftime("%H:%M"),
        'jornada_estandar': format_timedelta_to_hhmmss(JORNADA_ESTANDAR),
    }
    
    return render(request, 'control_horas/calculadora.html', context)


# =================================================================
# 6. VISTAS DE REPORTES (Exportación)
# =================================================================

@requiere_nivel(1)
def export_jornadas_csv(request):
    """
    Exporta los registros de jornada a un archivo CSV.
    """
    # 1. Preparar la respuesta HTTP para CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="reporte_jornadas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)

    # 2. Definir encabezados del CSV
    writer.writerow([
        'ID Registro',
        'Empleado (Username)',
        'Fecha',
        'Entrada',
        'Salida',
        'Horas Netas (HH:MM:SS)',
        'Horas Extra (HH:MM:SS)'
    ])

    # 3. Filtrado de datos (por parámetros GET)
    registros = RegistroJornada.objects.all().order_by('empleado__username', 'fecha')
    
    empleado_username = request.GET.get('empleado')
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    
    try:
        if empleado_username:
            registros = registros.filter(empleado__username=empleado_username)
        
        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            registros = registros.filter(fecha__gte=fecha_inicio)
            
        if fecha_fin_str:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            registros = registros.filter(fecha__lte=fecha_fin)
            
    except ValueError:
        messages.error(request, "Error en el formato de fecha de filtrado. Se esperaba AAAA-MM-DD.")
        return redirect('control_horas:saldo_horas') 


    # 4. Escribir los datos en el CSV
    for registro in registros:
        writer.writerow([
            registro.pk,
            registro.empleado.username,
            registro.fecha.isoformat(),
            registro.entrada.strftime('%H:%M:%S'),
            registro.salida.strftime('%H:%M:%S'),
            format_timedelta_to_hhmmss(registro.horas_netas),
            format_timedelta_to_hhmmss(registro.horas_extras),
        ])

    # 5. Escribir el contenido del buffer a la respuesta
    response.write(csv_buffer.getvalue())
    return response
def control_horas_dashboard(request):
    """
    Vista principal del dashboard de Control de Horas, dentro del módulo RR.HH.
    Muestra las opciones de gestión de registros de jornada, saldos y horas extras.
    """
    # En una aplicación real, aquí podrías pasar datos de resúmenes o KPIs 
    # de control de horas para mostrarlos en el dashboard.
    context = {
        'greeting': 'Bienvenido al Dashboard de Control de Horas.',
    }
    return render(request, 'control_horas/control_dashboard.html', context)